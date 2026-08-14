"""data_center_v2 - 多数据源管理。

DataSourceManager：统一接入 4 种数据源
  - CSV      （本地 csv 文件，含内置样例）
  - Excel    （.xlsx/.xls）
  - API      （官方开奖 API：webapi.sporttery.cn）
  - Database （SQLite/任意数据库连接串）

加载后生成 DataQualityReport 供 UI 判断数据可信度。
"""
from __future__ import annotations

import csv
import io
import json
import sqlite3
import time
import urllib.request
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from .models import DrawRecord, DataSourceInfo
from .quality import DataQualityReport

# 官方体彩 API（大乐透）
SPORTTERY_API = "https://webapi.sporttery.cn/gateway/lottery/getHistoryPageListV1.qry"
DLT_GAME_NO = "85"

# 官方福彩 API（双色球）。注意：双色球属福彩，不能用体彩 sporttery 的 gameNo=235。
CWL_API = "https://www.cwl.gov.cn/cwl_admin/front/cwlkj/search/kjxx/findDrawNotice"


def _parse_numbers(text: str, front_n: int, back_n: int) -> tuple[List[int], List[int]]:
    """'10 11 18 22 35|06 12' / '10 11 18 22 35 06 12' / '04,06,10,18,23,31|11' -> (front, back)。

    支持空格与逗号分隔。
    """
    text = text.replace(",", " ")
    if "|" in text:
        f_part, b_part = text.split("|", 1)
    else:
        parts = text.split()
        f_part = " ".join(parts[:front_n])
        b_part = " ".join(parts[front_n : front_n + back_n])
    return [int(x) for x in f_part.split()], [int(x) for x in b_part.split()]


class CSVDatasource:
    """CSV 数据源。"""

    type_name = "csv"

    def __init__(self, path: str, lottery: str = "dlt"):
        self.path = Path(path)
        self.lottery = lottery

    def load(self) -> List[DrawRecord]:
        if not self.path.exists():
            return []
        rows = []
        with open(self.path, newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                num = row.get("issue") or row.get("number") or ""
                date = row.get("date") or row.get("draw_date") or ""
                nums = row.get("numbers") or row.get("result") or ""
                pool = row.get("pool") or row.get("pool_balance") or ""
                if not num or not nums:
                    continue
                try:
                    front, back = _parse_numbers(nums, 5, 2)
                    if self.lottery == "ssq":
                        front, back = _parse_numbers(nums, 6, 1)
                except (ValueError, TypeError):
                    continue
                try:
                    pool_val = float(pool.replace(",", "")) if pool else 0.0
                except ValueError:
                    pool_val = 0.0
                rows.append(
                    DrawRecord(
                        number=num, draw_date=date,
                        front=front, back=back, pool=pool_val, lottery=self.lottery,
                    )
                )
        return rows


class ExcelDatasource:
    """Excel 数据源（.xlsx/.xls）。需 openpyxl。"""

    type_name = "excel"

    def __init__(self, path: str, lottery: str = "dlt", sheet: str = "Sheet1"):
        self.path = Path(path)
        self.lottery = lottery
        self.sheet = sheet

    def load(self) -> List[DrawRecord]:
        if not self.path.exists():
            return []
        try:
            from openpyxl import load_workbook
        except ImportError:
            return []
        wb = load_workbook(self.path, read_only=True, data_only=True)
        ws = wb[self.sheet] if self.sheet in wb.sheetnames else wb.active
        rows: List[DrawRecord] = []
        headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(max_row=1))]
        for r in ws.iter_rows(min_row=2, values_only=True):
            d = dict(zip(headers, r))
            num = str(d.get("issue", "") or d.get("number", "") or "")
            date = str(d.get("date", "") or d.get("draw_date", "") or "")
            nums = str(d.get("numbers", "") or d.get("result", "") or "")
            if not num or not nums:
                continue
            try:
                front, back = _parse_numbers(nums, 5, 2)
            except (ValueError, TypeError):
                continue
            try:
                pool = float(str(d.get("pool", 0)).replace(",", ""))
            except ValueError:
                pool = 0.0
            rows.append(DrawRecord(num, date, front, back, pool, self.lottery))
        wb.close()
        return rows


class APIDatasource:
    """官方开奖 API 数据源（体彩 webapi.sporttery.cn）。

    支持多彩种：大乐透 gameNo=85（5+2），双色球 gameNo=235（6+1）。
    """

    type_name = "api"

    # 彩种 -> 官方 gameNo（仅体彩；双色球属福彩，走 CWLDatasource）
    GAME_NOS = {"dlt": "85"}

    def __init__(self, lottery: str = "dlt", pages: int = 18, page_size: int = 30):
        self.lottery = lottery
        self.pages = pages
        self.page_size = page_size

    @property
    def _game_no(self) -> str:
        return self.GAME_NOS.get(self.lottery, "85")

    def _fetch_page(self, page_no: int) -> List[DrawRecord]:
        url = (
            f"{SPORTTERY_API}?gameNo={self._game_no}&provinceId=0"
            f"&pageSize={self.page_size}&isVerify=1&pageNo={page_no}"
        )
        req = urllib.request.Request(
            url, headers={"User-Agent": "Mozilla/5.0", "Referer": "https://static.sporttery.cn/"}
        )
        with urllib.request.urlopen(req, timeout=15) as r:
            data = json.loads(r.read().decode("utf-8"))
        out = []
        for item in (data.get("value") or {}).get("list") or []:
            num = item.get("lotteryDrawNum", "")
            result = item.get("lotteryDrawResult", "")
            dt = item.get("lotteryDrawTime", "")
            pool = item.get("poolBalanceAfterdraw", "")
            if not num or not result:
                continue
            if self.lottery == "ssq":
                front, back = _parse_numbers(result, 6, 1)
            else:
                front, back = _parse_numbers(result, 5, 2)
            try:
                pool_val = float(pool.replace(",", "")) if pool else 0.0
            except ValueError:
                pool_val = 0.0
            out.append(DrawRecord(num, dt, front, back, pool_val, self.lottery))
        return out

    def load(self) -> List[DrawRecord]:
        rows: List[DrawRecord] = []
        for p in range(1, self.pages + 1):
            try:
                rows.extend(self._fetch_page(p))
            except Exception:
                break
        return rows


class CWLDatasource:
    """官方福彩 API 数据源（双色球 ssq，v4.10.1 修复）。

    双色球属于中国福利彩票，之前误用体彩 sporttery 的 gameNo=235（返回空），
    导致双色球数据长期滞后。此源改调福彩官网公开 API。
    """

    type_name = "api"

    def __init__(self, lottery: str = "ssq", issue_count: int = 100):
        self.lottery = lottery
        self.issue_count = issue_count

    def load(self) -> List[DrawRecord]:
        url = f"{CWL_API}?name=ssq&issueCount={self.issue_count}"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        try:
            with urllib.request.urlopen(req, timeout=15) as r:
                data = json.loads(r.read().decode("utf-8"))
        except Exception:
            return []
        out: List[DrawRecord] = []
        for item in (data.get("result") or []):
            num = item.get("code", "")
            dt = item.get("date", "")
            red = item.get("red", "")
            blue = item.get("blue", "")
            if not num or not red:
                continue
            # 福彩返回 red="05,08,15,20,21,24"、blue="09"；date 形如 "2026-08-13(四)"
            front = [int(x) for x in red.replace(",", " ").split()]
            back = [int(x) for x in blue.replace(",", " ").split()] if blue else []
            if "(" in dt:
                dt = dt.split("(")[0]
            out.append(DrawRecord(num, dt, front, back, 0.0, self.lottery))
        # 福彩 API 返回从新到旧，反转为从旧到新（与其它数据源一致）
        out.reverse()
        return out


class DatabaseDatasource:
    """SQLite / 数据库数据源。"""

    type_name = "database"

    def __init__(self, conn_str: str, lottery: str = "dlt", table: str = "draws"):
        self.conn_str = conn_str
        self.lottery = lottery
        self.table = table

    def load(self) -> List[DrawRecord]:
        try:
            conn = sqlite3.connect(self.conn_str)
            cur = conn.execute(f"SELECT issue,date,numbers,pool FROM {self.table} ORDER BY issue DESC")
            rows = []
            for num, date, nums, pool in cur.fetchall():
                if not num or not nums:
                    continue
                try:
                    front, back = _parse_numbers(str(nums), 5, 2)
                except (ValueError, TypeError):
                    continue
                rows.append(DrawRecord(str(num), str(date), front, back, float(pool or 0), self.lottery))
            conn.close()
            return rows
        except (sqlite3.Error, Exception):
            return []


# 模块级缓存（v4.9.1 性能修复：避免每次匹配开奖都重复读 CSV 解析上千期）
# key: (lottery, base_dir) → (timestamp, List[DrawRecord])
_SOURCE_CACHE: Dict[Tuple[str, Optional[str]], Tuple[float, List]] = {}
_SOURCE_CACHE_TTL = 300  # 秒（5 分钟；数据更新后 TTL 过期自动重读）


class DataSourceManager:
    """数据源管理器：按优先级加载数据并给出质量报告。"""

    def __init__(self, lottery: str = "dlt"):
        self.lottery = lottery
        self._sources: List[object] = []
        self.draws: List[DrawRecord] = []
        self._cached_draws: Optional[List[DrawRecord]] = None
        self.report: Optional[DataQualityReport] = None
        self.info: Optional[DataSourceInfo] = None

    # ---- 源注册 ----
    def add_csv(self, path: str) -> "DataSourceManager":
        self._sources.append(CSVDatasource(path, self.lottery))
        return self

    def add_excel(self, path: str) -> "DataSourceManager":
        self._sources.append(ExcelDatasource(path, self.lottery))
        return self

    def add_api(self, pages: int = 18) -> "DataSourceManager":
        if self.lottery == "ssq":
            self._sources.append(CWLDatasource(self.lottery))
        else:
            self._sources.append(APIDatasource(self.lottery, pages=pages))
        return self

    def add_database(self, conn_str: str) -> "DataSourceManager":
        self._sources.append(DatabaseDatasource(conn_str, self.lottery))
        return self

    # ---- 加载 ----
    def load(self) -> List[DrawRecord]:
        """按注册顺序加载，取第一个有数据的源；数据不足时可合并。"""
        # 缓存命中：直接返回（避免重复读文件，性能优化 v4.9.1）
        if self._cached_draws is not None:
            self.draws = self._cached_draws
            return self.draws
        combined: List[DrawRecord] = []
        used: Optional[object] = None
        for src in self._sources:
            try:
                rows = src.load()
            except Exception:
                rows = []
            if rows:
                combined = rows
                used = src
                break  # 优先第一个有数据的源（用户数据目录优先）
        # 统一按时序升序排序（旧 → 新）
        combined = sorted(combined, key=lambda d: d.number)
        self.draws = combined
        self.report = DataQualityReport.build(
            self.lottery, combined,
            source_type=used.type_name if used else "none",
            source_path=str(getattr(used, "path", "")) if used else "",
        )
        self.info = DataSourceInfo(
            source_type=self.report.source_type,
            source_path=self.report.source_path,
            record_count=self.report.total,
            is_builtin=self.report.source_type in ("none",),
        )
        return self.draws

    def quality(self) -> DataQualityReport:
        if self.report is None:
            self.load()
        return self.report

    def as_plain_records(self) -> List[dict]:
        """转普通 dict，便于序列化/测试。"""
        return [
            {"number": d.number, "date": d.draw_date, "front": d.front, "back": d.back, "pool": d.pool}
            for d in self.draws
        ]

    @classmethod
    def from_project(cls, lottery: str = "dlt", base_dir: Optional[str] = None) -> "DataSourceManager":
        """从项目数据目录构建（用户数据优先，内置回退）。

        v4.9.1 修复：优先读用户增量缓存 `~/.atlas/raw/`（含最新开奖，如 26090），
        其次 `~/.atlas/data/`（旧版），最后项目内置。修复中奖计算漏算最新期的问题。
        同时加模块级缓存（TTL 5 分钟）避免每次匹配重复读 CSV 解析上千期。
        """
        key = (lottery, base_dir)
        now = time.time()
        cached = _SOURCE_CACHE.get(key)
        if cached and now - cached[0] < _SOURCE_CACHE_TTL:
            mgr = cls(lottery)
            mgr._cached_draws = cached[1]
            return mgr

        mgr = cls(lottery)
        if base_dir:
            raw = Path(base_dir) / "data" / "raw"
        else:
            raw = Path(__file__).resolve().parent.parent.parent / "data" / "raw"
        # 用户数据目录优先（raw = v4.3.1 增量更新器缓存，data = 旧版）
        user_home = Path.home() / ".atlas"
        for sub in ("raw", "data"):
            user_file = user_home / sub / f"{lottery}_history.csv"
            if user_file.exists():
                mgr.add_csv(str(user_file))
                mgr.load()
                _SOURCE_CACHE[key] = (now, mgr.draws)
                return mgr
        # 项目 data/raw（真实历史数据或样例）
        cands = [raw / f"{lottery}_history.csv", raw / f"{lottery}_2024_sample.csv"]
        for c in cands:
            if c.exists():
                mgr.add_csv(str(c))
                mgr.load()
                _SOURCE_CACHE[key] = (now, mgr.draws)
                break
        return mgr
